"""
偶极-偶极装置 xlsx -> RES2DINV .dat 转换脚本

数据来源: 常规电法偶极偶极数据.xlsx
输出: 偶极-偶极装置测深.dat (RES2DINV 格式)

工作参数 (来自 PDF):
  AB = MN = a = 4 m (偶极长度)
  n = 1, 2, 3 (隔离系数)
  点距 = 2 m (电极间距)
  装置类型 = 3 (dipole-dipole)

数据列:
  OO' : 记录点 x 位置 (AB中点 与 MN中点 连线中点)
  Ps  : 视电阻率 ρs (Ω·m)
  V   : 电位差 (mV)
  Fs  : 视幅频率 (不用, IP=0)
  I   : 电流 (mA)

几何关系 (MN 在 AB 右侧):
  记录点 x = AB中点 + a*(1+n)/2 = AB中点 + 2*(1+n)
  n=1 -> x = AB中点 + 4
  n=2 -> x = AB中点 + 6
  n=3 -> x = AB中点 + 8
  同组内 n=1,2,3 的记录点相差 2m (与数据吻合)

每 3 行一组 (同电流 I), 组内 3 行对应 n=1,2,3。
按组内 OO' 大小赋 n: 最小->n=1, 中->n=2, 最大->n=3。
"""
import openpyxl
import os

# ============ 路径 ============
SRC_XLSX = r"C:\Users\BBG\Desktop\实习地质资料\常规电法-2026年度\常规电法偶极偶极数据.xlsx"
OUT_DAT = r"C:\Users\BBG\Desktop\实习地质资料\常规电法-2026年度\偶极-偶极装置测深.dat"

# ============ 参数 ============
A_DIPOLE = 4          # 偶极长度 a = 4m (AB=MN=4m)
ELEC_SPACING = 2      # 电极间距 (点距) = 2m
ARRAY_TYPE = 3        # RES2DINV 装置类型: 3 = dipole-dipole
IP_FLAG = 0           # 0 = 纯电阻率数据 (无 IP)
LINE_TITLE = "Yueyang Dipole-Dipole Survey Line"

# 已知异常值修正 (行号从1开始, 含表头; 数据从第2行开始)
# 这些 OO' 值明显偏离同组等差 {m-2, m, m+2} 规律, 根据相邻组推断修正
ANOMALY_FIX = {
    # Row60: OO'=39, 同组应为 {32,34,36}, 39->34
    60: 34,
    # Row81: OO'=70, 同组应为 {18,20,22}, 70->20
    81: 20,
    # Row86: OO'=19, 同组应为 {14,16,18}, 19->14
    86: 14,
}


def read_xlsx(path):
    """读取 xlsx, 返回 [(oo, ps, v, fs, i), ...] (跳过表头)"""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Sheet1']
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row,
                                         values_only=True), start=2):
        oo, ps, v, fs, curr = row
        # 修正异常 OO' 值
        if i in ANOMALY_FIX:
            print(f"  [修正] 第{i}行 OO'={oo} -> {ANOMALY_FIX[i]}")
            oo = ANOMALY_FIX[i]
        rows.append((float(oo), float(ps), float(v), float(fs), float(curr)))
    return rows


def assign_n_values(rows):
    """
    每 3 行一组, 组内按 OO' 排序赋 n=1,2,3。
    返回 [(x, a, n, ps), ...]
    """
    assert len(rows) % 3 == 0, f"数据点数 {len(rows)} 不是 3 的倍数"
    n_groups = len(rows) // 3
    result = []

    for g in range(n_groups):
        group = rows[g*3:(g+1)*3]
        # 按 OO' 排序
        sorted_group = sorted(group, key=lambda r: r[0])
        for n, (oo, ps, v, fs, curr) in enumerate(sorted_group, start=1):
            result.append((oo, A_DIPOLE, n, ps))

        # 验证组内 OO' 等差
        oos = [r[0] for r in sorted_group]
        d1 = oos[1] - oos[0]
        d2 = oos[2] - oos[1]
        if abs(d1 - 2.0) > 0.01 or abs(d2 - 2.0) > 0.01:
            print(f"  [警告] 第{g+1}组 OO' 非等差2: {oos} (差 {d1}, {d2})")

    return result, n_groups


def write_dat(path, data, n_groups):
    """写 RES2DINV .dat 文件"""
    n_data = len(data)
    # 第5行: 装置中点位置 (0 = x 为绝对坐标, 沿测线距离)
    mid_flag = 0

    with open(path, 'w', encoding='ascii') as f:
        f.write(f"{LINE_TITLE}\n")
        f.write(f"{ELEC_SPACING}\n")
        f.write(f"{ARRAY_TYPE}\n")
        f.write(f"{n_data}\n")
        f.write(f"{mid_flag}\n")
        f.write(f"{IP_FLAG}\n")
        for x, a, n, ps in data:
            f.write(f"{x:.1f} {a} {n} {ps:.1f}\n")

    print(f"\n已写入: {path}")
    print(f"  数据点数: {n_data} ({n_groups} 组 x 3)")


def main():
    print("=" * 60)
    print("偶极-偶极装置 xlsx -> RES2DINV .dat 转换")
    print("=" * 60)

    print(f"\n[1] 读取 xlsx: {SRC_XLSX}")
    rows = read_xlsx(SRC_XLSX)
    print(f"  共 {len(rows)} 个数据点")

    print(f"\n[2] 分组赋 n 值 (每3行一组, n=1,2,3)")
    data, n_groups = assign_n_values(rows)

    # 统计
    xs = [d[0] for d in data]
    pss = [d[3] for d in data]
    print(f"  组数: {n_groups}")
    print(f"  x 范围: [{min(xs):.1f}, {max(xs):.1f}] m")
    print(f"  ρs 范围: [{min(pss):.1f}, {max(pss):.1f}] Ω·m")

    # n 分布
    for n in [1, 2, 3]:
        cnt = sum(1 for d in data if d[2] == n)
        print(f"  n={n}: {cnt} 个点")

    print(f"\n[3] 写 RES2DINV .dat")
    write_dat(OUT_DAT, data, n_groups)

    # 打印前 9 行预览
    print(f"\n[预览] 前 9 行数据:")
    print(f"  {'x':>6} {'a':>3} {'n':>3} {'ρs':>8}")
    for x, a, n, ps in data[:9]:
        print(f"  {x:6.1f} {a:3d} {n:3d} {ps:8.1f}")

    print("\n完成!")


if __name__ == '__main__':
    main()
